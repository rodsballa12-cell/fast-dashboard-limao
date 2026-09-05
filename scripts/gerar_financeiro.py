# -*- coding: utf-8 -*-
"""Le a aba DRE do painel Excel e gera data/financeiro.json com TODOS os meses.

Rodar depois de qualquer atualização do painel Excel na máquina do Rodrigo.
O Excel precisa ter sido aberto e salvo pelo Excel antes, senão as fórmulas
não têm valor em cache.

Estrutura esperada da aba DRE:
  - Coluna A: rótulos (categorias como "Comissão sobre produção", "Aluguel", etc)
  - Colunas subsequentes: um mês por coluna
  - Cabeçalho de mês pode estar em qualquer uma das primeiras 5 linhas
  - Formato aceito: "Jan/26", "Janeiro", "01/2026", datetime, "jan-26" etc

O script detecta automaticamente onde está o cabeçalho e quais colunas
são meses. Se não achar padrão, imprime diagnóstico claro.

Projetado × Realizado: o Excel mantém o projetado no mês corrente até
Rodrigo começar a atualizar com o realizado (a cada 15 dias). Nada muda
aqui — o valor lido é o que estiver na célula, seja projetado ou real.
"""
import json, os, datetime, re, unicodedata, sys
import openpyxl

PAINEL = r"C:\Users\rods_\OneDrive\Franquia - FAST\Claude\Painel_Gestao_Financeira_SIIBELLO.xlsx"
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "financeiro.json")

META_MES = 60000.00
PREM = dict(comissao=0.32, insumos=0.06, aluguel=18000.00, gerente=6500.00,
            midia=2500.00, beleza_boost=1900.00, sistemas=400.00,
            royalty_min=2500.00, mkt_local=1000.00)

MES_NUM = {
    "jan":1, "janeiro":1, "fev":2, "fevereiro":2, "mar":3, "marco":3, "março":3,
    "abr":4, "abril":4, "mai":5, "maio":5, "jun":6, "junho":6,
    "jul":7, "julho":7, "ago":8, "agosto":8, "set":9, "setembro":9,
    "out":10, "outubro":10, "nov":11, "novembro":11, "dez":12, "dezembro":12,
}


def _norm(x):
    t = unicodedata.normalize("NFKD", str(x or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", " ", t.lower()).strip()


def detectar_meses(ws, ano_default=None):
    """Varre as primeiras 5 linhas × 20 colunas procurando cabeçalhos de mês.
    Retorna dict {col: (ano, mes)} com colunas identificadas."""
    if ano_default is None:
        ano_default = datetime.date.today().year
    achados = {}
    for r in range(1, 6):
        for c in range(2, 21):
            v = ws.cell(r, c).value
            if v is None: continue
            # datetime direto
            if isinstance(v, datetime.datetime):
                achados[c] = (v.year, v.month); continue
            if isinstance(v, datetime.date):
                achados[c] = (v.year, v.month); continue
            s = str(v).lower().strip()
            # padrão "01/2026", "01/26"
            m = re.match(r"^(\d{1,2})[/\-](\d{2,4})$", s)
            if m:
                mes = int(m.group(1)); ano = int(m.group(2))
                if ano < 100: ano += 2000
                if 1 <= mes <= 12:
                    achados[c] = (ano, mes); continue
            # padrão "jan/26", "jan-26", "janeiro/2026", "Jan/26"
            m = re.match(r"^([a-z]{3,10})[/\-]?(\d{2,4})?$", s)
            if m:
                mes_nome = m.group(1)
                mes_num = MES_NUM.get(mes_nome)
                if mes_num:
                    ano = m.group(2)
                    ano = (int(ano) if ano else ano_default)
                    if ano < 100: ano += 2000
                    achados[c] = (ano, mes_num); continue
    # Retorna só se achou uma sequência plausível (>= 3 meses)
    if len(achados) >= 3:
        return achados
    return {}


def ler_dre(ws, col_mes):
    """Le a aba DRE por rótulo, retornando dict {rotulo_normalizado: valor}
    da coluna do mês específico. Valores None viram 0.0."""
    m = {}
    for r in range(1, ws.max_row + 1):
        rot = ws.cell(r, 1).value
        val = ws.cell(r, col_mes).value
        if rot:
            k = _norm(rot)
            if k and isinstance(val, (int, float)):
                m.setdefault(k, val)
    return m


def pegar(m, rotulo, obrigatorio=True):
    """Busca valor pelo rótulo, tratando sinal (custo vem negativo, receita positivo)."""
    k = _norm(rotulo)
    if k in m:
        return m[k]
    if obrigatorio:
        return None
    return 0.0


def monta_mes(dre_map, ano, mes):
    """Constrói o dict de um mês a partir dos valores lidos da DRE."""
    rec = pegar(dre_map, "Receita faturada (R$)") or pegar(dre_map, "Receita") or 0.0
    # custos vêm negativos no P&L; converte pra positivo
    def _custo(rot):
        v = pegar(dre_map, rot, obrigatorio=False)
        return -v if v and v < 0 else (v or 0.0)

    comissao = _custo("Comissão sobre produção — VARIÁVEL") or _custo("Comissão sobre produção")
    insumos = _custo("Produtos e insumos")
    gerente = _custo("Gerente — fixo mensal") or _custo("Gerente")
    clt_rec = _custo("Folha CLT — recepção (2 pessoas)") or _custo("Folha CLT recepção")
    clt_lim = _custo("Folha CLT — limpeza") or _custo("Folha CLT limpeza")
    vt = _custo("Vale-transporte da equipe") or _custo("Vale-transporte")
    aluguel = _custo("Aluguel + IPTU")
    marketing = _custo("Marketing (mídia, Beleza Boost, impressos, audiovisual)") or _custo("Marketing")
    utilidades = _custo("Utilidades, telecom e consumo")
    seguro = _custo("Seguro do imóvel")
    franqueadora = _custo("Franqueadora, sistemas e taxas") or _custo("Franqueadora")
    a_classificar = _custo("Outros ainda a classificar") or 0.0

    SIMPLES, INSUMOS_PCT = 0.07, PREM["insumos"]
    grupos = [
        ("VARIAVEL", "Custos variáveis — acompanham a venda", [
            ("Comissão sobre produção", comissao, PREM["comissao"] * rec,
             "premissa 32% da receita", False, {"esp_pct": PREM["comissao"], "real_delta_pct": comissao / max(rec,1)}),
            ("Produtos e insumos (CMV)", insumos, INSUMOS_PCT * rec,
             "premissa 6% da receita", False, {"esp_pct": INSUMOS_PCT, "real_delta_pct": insumos / max(rec,1)}),
            ("Impostos sobre venda — Simples", SIMPLES * rec, SIMPLES * rec,
             "7% da receita · provisionado por competência", True, {"esp_pct": SIMPLES, "real_pct": SIMPLES}),
        ]),
        ("PESSOAL", "Pessoal fixo — independe do faturamento", [
            ("Gerente", gerente, PREM["gerente"], "R$ 6.500/mês"),
            ("Folha CLT — recepção (2 pessoas)", clt_rec, 0.0, "não estava no modelo"),
            ("Folha CLT — limpeza", clt_lim, 0.0, "não estava no modelo"),
            ("Vale-transporte", vt, 0.0, "não estava no modelo"),
        ]),
        ("OCUPACAO", "Ocupação — o imóvel das duas lojas", [
            ("Aluguel + IPTU", aluguel, PREM["aluguel"], "R$ 18.000/mês na premissa"),
            ("Utilidades, telecom e consumo", utilidades, None, "sem premissa"),
            ("Seguro do imóvel", seguro, None, "sem premissa"),
        ]),
        ("COMERCIAL", "Comercial e franquia", [
            ("Marketing", marketing, PREM["midia"] + PREM["beleza_boost"], ""),
            ("Franqueadora, sistemas e taxas", franqueadora, PREM["sistemas"], "Trinks + Sults"),
            ("Royalty + marketing local", PREM["royalty_min"] + PREM["mkt_local"],
             PREM["royalty_min"] + PREM["mkt_local"], "mínimo contratual · provisionado", True),
        ]),
        ("ABERTO", "Ainda sem classificação", [
            ("Outros a classificar", a_classificar, 0.0, ""),
        ]),
    ]

    def soma(g, i):
        return sum((l[i] or 0.0) for l in g[2])

    G = [{"id": g[0], "titulo": g[1],
          "linhas": [dict({"cat": l[0], "real": l[1], "esp": l[2], "nota": l[3],
                           "prov": len(l) > 4 and l[4]}, **(l[5] if len(l) > 5 else {}))
                     for l in g[2]],
          "sub_real": soma(g, 1), "sub_esp": soma(g, 2)} for g in grupos]
    gi = {g["id"]: g for g in G}

    var_real = gi["VARIAVEL"]["sub_real"]; var_esp = gi["VARIAVEL"]["sub_esp"]
    fixo_real = sum(gi[k]["sub_real"] for k in ("PESSOAL", "OCUPACAO", "COMERCIAL", "ABERTO"))
    fixo_esp = sum(gi[k]["sub_esp"] for k in ("PESSOAL", "OCUPACAO", "COMERCIAL", "ABERTO"))
    real_total = var_real + fixo_real
    esp_total = var_esp + fixo_esp
    prov = sum(l["real"] for g in G for l in g["linhas"] if l["prov"])
    resultado = rec - real_total
    caixa_real = resultado + prov

    return {
        "ano": ano, "mes": mes, "chave": f"{ano}-{mes:02d}",
        "receita_real": rec, "receita_meta": META_MES,
        "grupos": G,
        "var_real": var_real, "var_esp": var_esp,
        "fixo_real": fixo_real, "fixo_esp": fixo_esp,
        "custo_real": real_total, "custo_esp": esp_total,
        "res_real": resultado, "res_caixa": caixa_real,
        "provisoes_nao_debitadas": prov,
    }


def main():
    if not os.path.exists(PAINEL):
        print(f"❌ Painel não encontrado: {PAINEL}"); sys.exit(1)
    wb = openpyxl.load_workbook(PAINEL, data_only=True)

    # 1. Aba DRE (com fallback pra P&L_Operacional se DRE não existir)
    if "DRE" in wb.sheetnames:
        ws_dre = wb["DRE"]
        print(f"[dre] usando aba 'DRE' · {ws_dre.max_row} linhas × {ws_dre.max_column} cols")
    elif "P&L_Operacional" in wb.sheetnames:
        ws_dre = wb["P&L_Operacional"]
        print("[dre] aba 'DRE' não encontrada · fallback pra 'P&L_Operacional'")
    else:
        print(f"❌ Nem 'DRE' nem 'P&L_Operacional' existem. Abas: {wb.sheetnames}"); sys.exit(1)

    # 2. Detectar colunas dos meses
    cols_meses = detectar_meses(ws_dre)
    if not cols_meses:
        print("❌ Não achei cabeçalhos de mês nas primeiras 5 linhas × 20 colunas.")
        print("   Formatos aceitos: 'Jan/26', 'Janeiro', '01/2026', datetime, etc")
        print("   Verifique se a linha do cabeçalho tem os meses e roda de novo.")
        # Mostra o que achou pra debug
        print("\n   Preview das primeiras 3 linhas × 12 cols:")
        for r in range(1, 4):
            row_vals = [str(ws_dre.cell(r, c).value or "")[:15] for c in range(1, 13)]
            print(f"     R{r}: {' | '.join(row_vals)}")
        sys.exit(1)
    print(f"[dre] {len(cols_meses)} meses detectados:")
    for c in sorted(cols_meses):
        ano, mes = cols_meses[c]
        print(f"     col {c} = {ano}-{mes:02d}")

    # 3. Ler cada mês
    meses = {}
    for col, (ano, mes) in cols_meses.items():
        dre_map = ler_dre(ws_dre, col)
        if not dre_map:
            print(f"[warn] col {col} ({ano}-{mes:02d}): dre_map vazio, pulando")
            continue
        m = monta_mes(dre_map, ano, mes)
        meses[m["chave"]] = m
        if m["receita_real"] > 0:
            print(f"     {m['chave']}: receita R$ {m['receita_real']:,.2f} · resultado R$ {m['res_real']:,.2f}")

    if not meses:
        print("❌ Nenhum mês com dados válidos"); sys.exit(1)

    # 4. Determinar mês corrente e último mês fechado
    hoje = datetime.date.today()
    chave_hoje = f"{hoje.year}-{hoje.month:02d}"
    mes_corrente = meses.get(chave_hoje)
    # Último mês fechado: mês corrente-1 com dados válidos
    meses_com_dados = sorted([k for k in meses if meses[k]["receita_real"] > 0])
    mes_fechado_chave = None
    for k in reversed(meses_com_dados):
        if k < chave_hoje:
            mes_fechado_chave = k; break
    mes_fechado = meses.get(mes_fechado_chave) if mes_fechado_chave else None

    # Kpis: se tem mês corrente, usa. Senão, último fechado.
    principal = mes_corrente if mes_corrente and mes_corrente["receita_real"] > 0 else mes_fechado
    if not principal:
        # Fallback: primeiro mês com dados
        principal = meses[meses_com_dados[0]] if meses_com_dados else next(iter(meses.values()))

    # ---- Ponto de equilíbrio (herdado do modelo antigo) ----
    FIXO = 41816.32 + 1300.00  # CashMe/Pronampe fora
    rec_p = principal["receita_real"]
    com_p = principal["grupos"][0]["linhas"][0]["real"]
    com_pct = com_p / max(rec_p, 1)
    SIMPLES, INSUMOS_PCT = 0.07, PREM["insumos"]
    mc_meta = 1 - PREM["comissao"] - INSUMOS_PCT - SIMPLES
    fixo_modelo = PREM["aluguel"] + PREM["gerente"] + PREM["midia"] + PREM["beleza_boost"] \
                  + PREM["sistemas"] + PREM["royalty_min"] + PREM["mkt_local"]

    # ---- Payload ----
    d = {
        "gerado_em": datetime.datetime.now().astimezone().isoformat(timespec="seconds"),
        "baseline": hoje.isoformat(),
        "custos_ate": hoje.isoformat(),
        "mes_corrente_chave": mes_corrente["chave"] if mes_corrente else None,
        "mes_fechado_chave": mes_fechado_chave,
        "kpis": {
            "caixa_conta": 5060.93,   # TODO: puxar do Stone/extrato
            "a_receber_stone": 20742.90,
            "resultado_mes": principal["res_real"],
            "resultado_mes_caixa": principal["res_caixa"],
            "receita_mes": principal["receita_real"],
            "meta_mes": META_MES,
        },
        "meses": meses,  # todos os meses lidos da DRE
        "resultado": {  # mês PRINCIPAL (corrente com dados, ou último fechado)
            **principal,
            "res_esp_mesma_receita": rec_p - principal["custo_esp"],
            "res_esp_na_meta": META_MES * mc_meta - fixo_modelo,
            "be_modelo_1loja": fixo_modelo / mc_meta,
            "mc_real": rec_p - principal["var_real"],
            "mc_esp": rec_p - principal["var_esp"],
        },
        "premissas": {"comissao": PREM["comissao"], "insumos": INSUMOS_PCT, "simples": SIMPLES},
        "equilibrio": {
            "fatura_hoje": rec_p,
            "custo_fixo_mes": FIXO,
            "cenarios": [
                {"nome": "Com a comissão atual", "comissao": com_pct, "mc": 1 - com_pct - INSUMOS_PCT - SIMPLES},
                {"nome": "Se subir para 40%", "comissao": 0.40, "mc": 1 - 0.40 - INSUMOS_PCT - SIMPLES},
            ],
        },
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=1)
    print(f"\n✓ gerado: {OUT}")
    print(f"  {len(meses)} meses no JSON · mês principal: {principal['chave']}")


if __name__ == "__main__":
    main()
