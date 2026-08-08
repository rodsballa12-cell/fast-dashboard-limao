r"""Captura horaria (delta) — Trinks.

Roda de 1h em 1h. Puxa dados do dia corrente + agregados do mes,
sobrescreve dados/live/hoje_*.xlsx e regenera dashboard_data.json.

Custo por rodada: ~7 requisicoes (cabe folgado em 10.000/mes).

Uso:
    python trinks_delta.py                # dia de hoje
    python trinks_delta.py --mes-todo     # forca reprocessar o mes inteiro
"""
from __future__ import annotations

import json
import sys
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from trinks_common import TrinksClient

BASE = Path(__file__).resolve().parent.parent
DADOS = BASE / "dados"
LIVE = DADOS / "live"
HIST = LIVE / "historico"
LOGS = LIVE / "logs"
DASHBOARD_JSON = DADOS / "dashboard_data.json"

for p in (LIVE, HIST, LOGS):
    p.mkdir(parents=True, exist_ok=True)


def to_xlsx(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "dados"
    if not rows:
        ws.append(["(vazio)"])
        wb.save(out_path)
        return
    # achatar
    def flatten(d, prefix=""):
        out = {}
        if isinstance(d, dict):
            for k, v in d.items():
                key = f"{prefix}.{k}" if prefix else k
                if isinstance(v, dict):
                    out.update(flatten(v, key))
                elif isinstance(v, list):
                    out[key] = json.dumps(v, ensure_ascii=False, default=str)[:32000]
                else:
                    out[key] = v
        return out
    flat = [flatten(r) for r in rows]
    seen = {}
    for r in flat:
        for k in r:
            seen.setdefault(k, None)
    headers = list(seen.keys())
    ws.append(headers)
    for r in flat:
        ws.append([r.get(h, "") for h in headers])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(h) + 2, 12), 40)
    wb.save(out_path)


def read_xlsx(path: Path) -> tuple[list[str], list[tuple]]:
    if not path.exists():
        return [], []
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


def agregar_mes(agend_path: Path, transac_path: Path, vendas_path: Path, clientes_path: Path) -> dict:
    """Le os .xlsx do mes corrente e monta o resumo do dashboard."""
    h, d = read_xlsx(agend_path)
    fin, canc, total = [], [], 0
    days_series = []
    prof_list = []
    serv_list = []
    top_cli = []
    receita_serv = 0.0
    ticket = 0.0
    unicos = 0
    if h:
        i_status = h.index("status.nome") if "status.nome" in h else -1
        i_valor  = h.index("valor")       if "valor" in h else -1
        i_prof   = h.index("profissional.nome") if "profissional.nome" in h else -1
        i_serv   = h.index("servico.nome")      if "servico.nome" in h else -1
        i_data   = h.index("dataHoraInicio")    if "dataHoraInicio" in h else -1
        i_cli    = h.index("cliente.nome")      if "cliente.nome" in h else -1
        total = len(d)
        fin  = [r for r in d if i_status >= 0 and r[i_status] == "Finalizado"]
        canc = [r for r in d if i_status >= 0 and r[i_status] == "Cancelado"]

        by_day = defaultdict(lambda: {"n": 0, "v": 0.0})
        for r in fin:
            dt = r[i_data] if i_data >= 0 else None
            if isinstance(dt, str):
                try: dt = datetime.fromisoformat(dt.replace("Z", ""))
                except Exception: dt = None
            day = dt.day if dt else 0
            by_day[day]["n"] += 1
            by_day[day]["v"] += float(r[i_valor] or 0) if i_valor >= 0 else 0
        days_series = [{"d": k, "n": v["n"], "v": round(v["v"], 2)} for k, v in sorted(by_day.items())]

        prof_agg = defaultdict(lambda: {"n": 0, "v": 0.0})
        for r in fin:
            p = (r[i_prof] or "sem nome") if i_prof >= 0 else "sem nome"
            prof_agg[p]["n"] += 1
            prof_agg[p]["v"] += float(r[i_valor] or 0) if i_valor >= 0 else 0
        prof_list = sorted(
            [{"nome": k.strip().title(), "n": v["n"], "v": round(v["v"], 2)} for k, v in prof_agg.items()],
            key=lambda x: -x["v"],
        )

        serv_agg = defaultdict(lambda: {"n": 0, "v": 0.0})
        for r in fin:
            s = (r[i_serv] or "sem servico") if i_serv >= 0 else "sem servico"
            serv_agg[s]["n"] += 1
            serv_agg[s]["v"] += float(r[i_valor] or 0) if i_valor >= 0 else 0
        serv_list = sorted(
            [{"nome": k, "n": v["n"], "v": round(v["v"], 2)} for k, v in serv_agg.items()],
            key=lambda x: -x["v"],
        )[:10]

        cli_count = Counter(r[i_cli] for r in fin if i_cli >= 0 and r[i_cli])
        top_cli = [{"nome": k.title(), "n": v} for k, v in cli_count.most_common(10)]
        unicos = len(cli_count)
        if fin:
            receita_serv = round(sum(float(r[i_valor] or 0) for r in fin if i_valor >= 0), 2)
            ticket = round(receita_serv / len(fin), 2)

    # Transacoes (financeiro)
    # caixa_recebido = sum(formasPagamentos.valor) — bate com o relatorio Trinks (Data de Pagamento/Estorno)
    # totalPagar = valor liquido pos-desconto e pos-troco (nao bate com Trinks tela)
    h2, d2 = read_xlsx(transac_path)
    transacoes_total, descontos_total, caixa_recebido = 0.0, 0.0, 0.0
    fp_list = []
    if h2:
        i_tot = h2.index("totalPagar") if "totalPagar" in h2 else -1
        i_desc = h2.index("descontos") if "descontos" in h2 else -1
        i_fp = h2.index("formasPagamentos") if "formasPagamentos" in h2 else -1
        transacoes_total = round(sum(float(r[i_tot] or 0) for r in d2 if i_tot >= 0), 2)
        descontos_total = round(sum(float(r[i_desc] or 0) for r in d2 if i_desc >= 0), 2)
        fp_count = Counter()
        fp_valor = defaultdict(float)
        if i_fp >= 0:
            for r in d2:
                raw = r[i_fp]
                if not raw: continue
                try:
                    arr = json.loads(raw) if isinstance(raw, str) else raw
                    for it in arr:
                        nome = it.get("nome") or it.get("formaPagamento") or "outros"
                        v = float(it.get("valor") or 0)
                        # troco entra como negativo — inclui na soma para bater com Trinks
                        fp_count[nome] += 1
                        fp_valor[nome] += v
                        caixa_recebido += v
                except Exception:
                    pass
        fp_list = sorted(
            [{"nome": k, "n": fp_count[k], "v": round(fp_valor[k], 2)} for k in fp_count],
            key=lambda x: -x["v"],
        )
        caixa_recebido = round(caixa_recebido, 2)

    # Vendas
    h3, d3 = read_xlsx(vendas_path)
    vendas_total = 0.0
    if h3 and "valorLiquido" in h3:
        i_v = h3.index("valorLiquido")
        vendas_total = round(sum(float(r[i_v] or 0) for r in d3), 2)

    # Clientes base
    h4, d4 = read_xlsx(clientes_path)
    base_clientes = len(d4)

    return {
        "kpis": {
            "receita_servicos": receita_serv,
            "atendimentos": len(fin),
            "ticket_medio": ticket,
            "cancelamentos": len(canc),
            "taxa_cancelamento": round(len(canc) / max(total, 1) * 100, 1),
            "total_agendamentos": total,
            "clientes_unicos_atendidos": unicos,
            "base_clientes": base_clientes,
            "vendas_produtos": vendas_total,
            "transacoes_totalpagar": transacoes_total,
            "caixa_recebido": caixa_recebido,
            "descontos_totais": descontos_total,
            "diff_agend_vs_caixa": round(receita_serv - caixa_recebido, 2),
            "ticket_por_transacao_caixa": round(caixa_recebido / max(len([r for r in d2]), 1), 2) if h2 else 0,
        },
        "por_dia": days_series,
        "profissionais": prof_list,
        "servicos_top": serv_list,
        "clientes_top": top_cli,
        "formas_pagamento": fp_list,
    }


def main() -> None:
    forca_mes = "--mes-todo" in sys.argv
    cli = TrinksClient()
    q0 = cli.consumo()

    hoje = date.today()
    y, m = hoje.year, hoje.month
    ini_mes = f"{y:04d}-{m:02d}-01"
    fim_mes = f"{y:04d}-{m:02d}-{monthrange(y, m)[1]:02d}"
    hoje_str = hoje.isoformat()
    stamp = datetime.now().strftime("%Y-%m-%dT%H")

    # 1) captura o DIA de hoje (leve)
    dia_path_agend  = LIVE / f"hoje_agendamentos.xlsx"
    dia_path_trans  = LIVE / f"hoje_transacoes.xlsx"
    dia_path_vendas = LIVE / f"hoje_vendas.xlsx"

    print(f"[{stamp}] Delta run — hoje={hoje_str}  cota={q0.get('totalUtilizado')}/{q0.get('cotaTotal')}")
    dia_agend = list(cli.paginate("/v1/agendamentos", {"dataInicio": hoje_str, "dataFim": hoje_str}))
    to_xlsx(dia_agend, dia_path_agend)
    dia_trans = list(cli.paginate("/v1/transacoes", {"dataInicio": hoje_str, "dataFim": hoje_str}))
    to_xlsx(dia_trans, dia_path_trans)
    dia_vendas = list(cli.paginate("/v1/vendas", {"dataInicio": hoje_str, "dataFim": hoje_str}))
    to_xlsx(dia_vendas, dia_path_vendas)

    # 2) agregacao do MES CORRENTE: puxa o mes inteiro se --mes-todo ou se hoje = dia 1
    mes_agend_path  = DADOS / "agendamentos" / f"agendamentos_{y}-{m:02d}.xlsx"
    mes_trans_path  = DADOS / "financeiro"   / f"transacoes_{y}-{m:02d}.xlsx"
    mes_vendas_path = DADOS / "financeiro"   / f"vendas_{y}-{m:02d}.xlsx"
    mes_cli_path    = DADOS / "clientes"     / f"clientes_{y}-{m:02d}.xlsx"

    if forca_mes or not mes_agend_path.exists():
        print("  -> puxando MES CORRENTE completo")
        mes_agend = list(cli.paginate("/v1/agendamentos", {"dataInicio": ini_mes, "dataFim": fim_mes}))
        to_xlsx(mes_agend, mes_agend_path)
        mes_trans = list(cli.paginate("/v1/transacoes", {"dataInicio": ini_mes, "dataFim": fim_mes}))
        to_xlsx(mes_trans, mes_trans_path)
        mes_vendas = list(cli.paginate("/v1/vendas", {"dataInicio": ini_mes, "dataFim": fim_mes}))
        to_xlsx(mes_vendas, mes_vendas_path)
        mes_cli = list(cli.paginate("/v1/clientes", {}))
        to_xlsx(mes_cli, mes_cli_path)
    else:
        # merge do dia sobre o mes (regrava xlsx do mes) — leve
        print("  -> mesclando dia no snapshot do mes")
        for dia_p, mes_p in [(dia_path_agend, mes_agend_path),
                             (dia_path_trans, mes_trans_path),
                             (dia_path_vendas, mes_vendas_path)]:
            # 'id' e chave primaria em todos os 3 endpoints
            h_m, rows_m = read_xlsx(mes_p)
            h_d, rows_d = read_xlsx(dia_p)
            if not h_d: continue
            if not h_m: h_m = h_d
            # dict por id
            def to_dict(headers, rows):
                if "id" not in headers: return {}, headers
                i_id = headers.index("id")
                return {r[i_id]: dict(zip(headers, r)) for r in rows if r[i_id] is not None}, headers
            base, _ = to_dict(h_m, rows_m)
            novos, _ = to_dict(h_d, rows_d)
            base.update(novos)  # dia sobrescreve mes
            merged = list(base.values())
            # normaliza headers (uniao)
            all_h = list(h_m)
            for k in h_d:
                if k not in all_h: all_h.append(k)
            rows_out = [{k: r.get(k, "") for k in all_h} for r in merged]
            to_xlsx(rows_out, mes_p)

    # 3) agrega e escreve dashboard_data.json
    resumo = agregar_mes(mes_agend_path, mes_trans_path, mes_vendas_path, mes_cli_path)
    resumo["periodo"] = f"{y:04d}-{m:02d}"
    resumo["gerado_em"] = datetime.now().isoformat(timespec="seconds")
    q1 = cli.consumo()
    resumo["cota"] = {
        "plano": q1.get("plano"),
        "total": q1.get("cotaTotal"),
        "usadas": q1.get("totalUtilizado"),
        "saldo":  q1.get("saldoRestante"),
    }
    resumo["proxima_execucao_prevista"] = "de 1h em 1h (Task Scheduler)"
    DASHBOARD_JSON.write_text(json.dumps(resumo, ensure_ascii=False, indent=2), encoding="utf-8")
    (HIST / f"resumo_{stamp}.json").write_text(json.dumps(resumo, ensure_ascii=False), encoding="utf-8")

    log_line = (
        f"{datetime.now().isoformat(timespec='seconds')} "
        f"hoje_agend={len(dia_agend)} hoje_trans={len(dia_trans)} "
        f"cota={q1.get('totalUtilizado')}/{q1.get('cotaTotal')} "
        f"receita_mes={resumo['kpis']['receita_servicos']} "
        f"atendimentos_mes={resumo['kpis']['atendimentos']}\n"
    )
    (LOGS / f"delta_{y}-{m:02d}.log").open("a", encoding="utf-8").write(log_line)
    print(f"  -> {log_line.strip()}")
    print(f"  -> dashboard_data.json atualizado")


if __name__ == "__main__":
    main()
