r"""Gera dashboard_data.json COMPLETO — tudo que o index.html dinâmico precisa.

Estrutura do JSON:
  {
    "gerado_em": iso,
    "meta_mensal": 60000,
    "dias_op_mes": 26,
    "cota": {...},
    "abas": {
      "anual":   {kpis, categorias, top_clientes_ltv, rentabilidade_hora, hora_media, servicos_evol},
      "mensal":  {kpis, meta_progresso, categorias, novos_vs_recorr, rentabilidade_hora, dia_mes, hora_media, ranking_prof, ranking_serv},
      "semanal": {kpis, meta_progresso, categorias, dow, hora_media, ranking_prof, ranking_serv},
      "diario":  {kpis, meta_progresso, categorias, hora_abs, ranking_prof, ranking_serv, comparativo}
    }
  }

Roda depois do trinks_delta.py (que garante os .xlsx frescos do mês).
"""
from __future__ import annotations

import json
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
DADOS = BASE / "dados"
OUT_JSON = DADOS / "dashboard_data.json"
META_MENSAL = 60000
DIAS_OP_MES = 26  # exclui domingos
DOW_NOMES = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def load_xlsx(rel_path: str):
    fp = DADOS / rel_path
    if not fp.exists():
        return [], []
    wb = load_workbook(fp, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


def pd(v):
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v.replace("Z", ""))
        except Exception:
            return None
    if isinstance(v, datetime):
        return v
    return None


def parse_json_col(v):
    if not v:
        return []
    try:
        return json.loads(v) if isinstance(v, str) else v
    except Exception:
        return []


# ============================================================
# CARREGA TUDO — todos os meses disponíveis
# ============================================================

def descobrir_meses():
    """Lista todos os meses YYYY-MM com arquivos de agendamentos."""
    meses = set()
    for p in (DADOS / "agendamentos").glob("agendamentos_*.xlsx"):
        stem = p.stem.replace("agendamentos_", "")
        if len(stem) == 7 and stem[4] == "-":
            meses.add(stem)
    return sorted(meses)


def _normalizar_linhas(fichas):
    """fichas = lista de tuplas (header_local, rows_locais).
    Retorna (header_canonico, rows_reordenadas) — arquivos podem ter mesmas colunas em ordens diferentes."""
    if not fichas:
        return [], []
    # header canônico = união preservando ordem do primeiro arquivo
    canonico = list(fichas[0][0])
    for h_local, _ in fichas[1:]:
        for c in h_local:
            if c not in canonico:
                canonico.append(c)
    idx = {c: i for i, c in enumerate(canonico)}
    saida = []
    for h_local, rows_local in fichas:
        map_local_para_canon = [idx[c] for c in h_local]  # posição canônica de cada col local
        for r in rows_local:
            nova = [None] * len(canonico)
            for j, v in enumerate(r):
                if j < len(map_local_para_canon):
                    nova[map_local_para_canon[j]] = v
            saida.append(tuple(nova))
    return canonico, saida


def carregar_agendamentos():
    fichas = []
    for ym in descobrir_meses():
        h, d = load_xlsx(f"agendamentos/agendamentos_{ym}.xlsx")
        if h:
            fichas.append((h, d))
    return _normalizar_linhas(fichas)


def carregar_transacoes():
    fichas = []
    for ym in descobrir_meses():
        h, d = load_xlsx(f"financeiro/transacoes_{ym}.xlsx")
        if h:
            fichas.append((h, d))
    return _normalizar_linhas(fichas)


def carregar_clientes_cad():
    """Mapa cliente_id -> data_cadastro (para novos vs recorrentes)."""
    meses = descobrir_meses()
    mapa = {}
    for ym in meses:
        h, d = load_xlsx(f"clientes/clientes_{ym}.xlsx")
        if not h:
            continue
        i_id = h.index("id")
        i_dt = h.index("dataCadastro") if "dataCadastro" in h else -1
        for r in d:
            if r[i_id] is not None and i_dt >= 0:
                mapa[r[i_id]] = pd(r[i_dt])
    return mapa


# ============================================================
# CALCULOS
# ============================================================

def analisar_agendamentos(agend_rows, h, ini: date, fim: date):
    """KPIs + rankings + distribuições de um período."""
    i_st = h.index("status.nome")
    i_v = h.index("valor")
    i_p = h.index("profissional.nome")
    i_s = h.index("servico.nome")
    i_dt = h.index("dataHoraInicio")
    i_cl = h.index("cliente.id")
    i_cln = h.index("cliente.nome")
    i_dur = h.index("duracaoEmMinutos") if "duracaoEmMinutos" in h else -1

    todos = [r for r in agend_rows if pd(r[i_dt]) and ini <= pd(r[i_dt]).date() <= fim]
    fin = [r for r in todos if r[i_st] == "Finalizado"]
    canc = [r for r in todos if r[i_st] == "Cancelado"]

    receita = sum(float(r[i_v] or 0) for r in fin)
    unicos = len({r[i_cl] for r in fin if r[i_cl]})

    dias_com_op = len({pd(r[i_dt]).date() for r in fin if pd(r[i_dt])})

    # rankings
    prof = defaultdict(lambda: {"n": 0, "v": 0.0})
    for r in fin:
        p = (r[i_p] or "").strip().title() or "—"
        prof[p]["n"] += 1
        prof[p]["v"] += float(r[i_v] or 0)
    top_prof = sorted(
        [{"nome": k, "n": v["n"], "v": round(v["v"], 2)} for k, v in prof.items()],
        key=lambda x: -x["v"],
    )[:12]

    serv = defaultdict(lambda: {"n": 0, "v": 0.0, "min": 0})
    for r in fin:
        s = r[i_s] or "sem"
        serv[s]["n"] += 1
        serv[s]["v"] += float(r[i_v] or 0)
        if i_dur >= 0:
            serv[s]["min"] += int(r[i_dur] or 0)
    top_serv = sorted(
        [{"nome": k, "n": v["n"], "v": round(v["v"], 2), "min": v["min"]} for k, v in serv.items()],
        key=lambda x: -x["v"],
    )[:12]

    # rentabilidade R$/hora por serviço (só os com duração>0)
    rent_hora = []
    for k, v in serv.items():
        if v["min"] > 0:
            rent_hora.append({
                "nome": k,
                "n": v["n"],
                "v": round(v["v"], 2),
                "min_medio": round(v["min"] / max(v["n"], 1)),
                "ticket": round(v["v"] / max(v["n"], 1), 2),
                "rs_hora": round(v["v"] / v["min"] * 60, 2),
            })
    rent_hora.sort(key=lambda x: -x["rs_hora"])

    # por dia da semana (usa fin)
    by_dow = defaultdict(lambda: {"n": 0, "v": 0.0})
    for r in fin:
        dt = pd(r[i_dt])
        if dt:
            by_dow[DOW_NOMES[dt.weekday()]]["n"] += 1
            by_dow[DOW_NOMES[dt.weekday()]]["v"] += float(r[i_v] or 0)
    dow_list = [
        {"nome": n, "n": by_dow[n]["n"], "v": round(by_dow[n]["v"], 2)} for n in DOW_NOMES
    ]

    # por dia do mês
    by_day = defaultdict(lambda: {"n": 0, "v": 0.0})
    for r in fin:
        dt = pd(r[i_dt])
        if dt:
            by_day[dt.day]["n"] += 1
            by_day[dt.day]["v"] += float(r[i_v] or 0)
    dia_list = [{"d": d, "n": by_day[d]["n"], "v": round(by_day[d]["v"], 2)} for d in sorted(by_day)]

    # clientes recorrentes (top por visitas)
    cli_c = Counter(r[i_cln] for r in fin if r[i_cln])
    top_cli = [{"nome": (k or "").title(), "n": v} for k, v in cli_c.most_common(10)]

    return {
        "receita_serv": round(receita, 2),
        "atend_fin": len(fin),
        "atend_canc": len(canc),
        "atend_total": len(todos),
        "taxa_canc": round(len(canc) / max(len(todos), 1) * 100, 1),
        "clientes_unicos": unicos,
        "ticket_atend": round(receita / max(len(fin), 1), 2),
        "dias_com_op": dias_com_op,
        "ranking_prof": top_prof,
        "ranking_serv": top_serv,
        "rentabilidade_hora": rent_hora,
        "por_dow": dow_list,
        "por_dia_mes": dia_list,
        "clientes_top": top_cli,
    }


def analisar_transacoes(trans_rows, h, ini: date, fim: date):
    """Categorias + caixa + hora + meios pagamento."""
    i_dt = h.index("dataHora")
    i_fp = h.index("formasPagamentos")
    i_pac = h.index("pacotes")
    i_prod = h.index("produtos")
    i_serv = h.index("servicos")
    i_desc = h.index("descontos")
    i_troco = h.index("troco")
    i_tot = h.index("totalPagar")

    filt = [r for r in trans_rows if pd(r[i_dt]) and ini <= pd(r[i_dt]).date() <= fim]

    caixa = 0.0
    pac_v = 0.0
    pac_n = 0
    prod_v = 0.0
    prod_n = 0
    serv_v = 0.0
    serv_n = 0
    descontos = 0.0
    trocos = 0.0

    mp_c = Counter()
    mp_v = defaultdict(float)
    by_hora_c = defaultdict(int)
    by_hora_v = defaultdict(float)

    for r in filt:
        fp_arr = parse_json_col(r[i_fp])
        for it in fp_arr:
            v = float(it.get("valor") or 0)
            nome = it.get("nome") or "outros"
            caixa += v
            mp_c[nome] += 1
            mp_v[nome] += v

        for p in parse_json_col(r[i_pac]):
            q = int(p.get("quantidade") or 1)
            pac_v += float(p.get("valorUnitario") or 0) * q
            pac_n += q
        for p in parse_json_col(r[i_prod]):
            q = int(p.get("quantidade") or 1)
            prod_v += float(p.get("valorUnitario") or 0) * q
            prod_n += q
        for s in parse_json_col(r[i_serv]):
            serv_v += float(s.get("preco") or 0)
            serv_n += 1

        descontos += float(r[i_desc] or 0)
        trocos += float(r[i_troco] or 0)

        dt = pd(r[i_dt])
        v_tot = float(r[i_tot] or 0)
        if dt:
            by_hora_c[dt.hour] += 1
            by_hora_v[dt.hour] += v_tot

    meios = sorted(
        [{"nome": k, "n": mp_c[k], "v": round(mp_v[k], 2),
          "pct": round(mp_v[k] / max(caixa, 1) * 100, 1)} for k in mp_c],
        key=lambda x: -x["v"],
    )

    hora_abs = [{"h": h, "n": by_hora_c.get(h, 0), "v": round(by_hora_v.get(h, 0), 2)}
                for h in range(8, 21)]

    return {
        "caixa": round(caixa, 2),
        "n_trans": len(filt),
        "ticket_trans": round(caixa / max(len(filt), 1), 2),
        "descontos": round(descontos, 2),
        "trocos": round(trocos, 2),
        "categorias": {
            "pacotes":  {"v": round(pac_v, 2),  "n": pac_n,  "pct": round(pac_v / max(caixa, 1) * 100, 1)},
            "servicos": {"v": round(serv_v, 2), "n": serv_n, "pct": round(serv_v / max(caixa, 1) * 100, 1)},
            "produtos": {"v": round(prod_v, 2), "n": prod_n, "pct": round(prod_v / max(caixa, 1) * 100, 1)},
        },
        "meios_pagamento": meios,
        "hora_abs": hora_abs,
    }


def novos_vs_recorrentes(agend_fin_ago, h_ag, cadastro_map, ini_mes: date):
    i_v = h_ag.index("valor")
    i_cl = h_ag.index("cliente.id")

    novos_ids, recorr_ids = set(), set()
    for r in agend_fin_ago:
        cid = r[i_cl]
        dt_cad = cadastro_map.get(cid)
        if dt_cad is None:
            continue
        if dt_cad.date() >= ini_mes:
            novos_ids.add(cid)
        else:
            recorr_ids.add(cid)

    receita_novos = sum(float(r[i_v] or 0) for r in agend_fin_ago if r[i_cl] in novos_ids)
    receita_rec = sum(float(r[i_v] or 0) for r in agend_fin_ago if r[i_cl] in recorr_ids)
    atend_novos = sum(1 for r in agend_fin_ago if r[i_cl] in novos_ids)
    atend_rec = sum(1 for r in agend_fin_ago if r[i_cl] in recorr_ids)

    return {
        "novos": {"clientes": len(novos_ids), "atend": atend_novos, "receita": round(receita_novos, 2)},
        "recorrentes": {"clientes": len(recorr_ids), "atend": atend_rec, "receita": round(receita_rec, 2)},
    }


def top_ltv(agend_rows, h_ag, ini: date, fim: date, limite=15):
    i_st = h_ag.index("status.nome")
    i_v = h_ag.index("valor")
    i_cl = h_ag.index("cliente.id")
    i_cln = h_ag.index("cliente.nome")
    i_dt = h_ag.index("dataHoraInicio")

    ltv = defaultdict(lambda: {"n": 0, "v": 0.0, "nome": ""})
    for r in agend_rows:
        if r[i_st] != "Finalizado":
            continue
        dt = pd(r[i_dt])
        if not dt or not (ini <= dt.date() <= fim):
            continue
        cid = r[i_cl]
        ltv[cid]["n"] += 1
        ltv[cid]["v"] += float(r[i_v] or 0)
        ltv[cid]["nome"] = r[i_cln]

    lst = sorted(
        [{"nome": (d["nome"] or "").title(), "n": d["n"], "v": round(d["v"], 2)}
         for d in ltv.values()],
        key=lambda x: -x["v"],
    )
    total = sum(x["v"] for x in lst)
    uma_vez = sum(1 for x in lst if x["n"] == 1)
    p20_n = max(1, len(lst) * 20 // 100)
    p20_v = sum(x["v"] for x in lst[:p20_n])

    return {
        "total_clientes": len(lst),
        "receita_total": round(total, 2),
        "uma_vez": uma_vez,
        "duas_mais": len(lst) - uma_vez,
        "pareto20_v": round(p20_v, 2),
        "pareto20_pct": round(p20_v / max(total, 1) * 100, 1),
        "pareto20_n": p20_n,
        "top": lst[:limite],
    }


def calc_meta(caixa: float, meta: float, dias_realizados: int, dias_total: int):
    pct = caixa / max(meta, 1) * 100
    falta = meta - caixa
    dias_restantes = max(dias_total - dias_realizados, 0)
    necessario = falta / max(dias_restantes, 1) if dias_restantes else 0
    ritmo = caixa / max(dias_realizados, 1)
    projecao = ritmo * dias_total
    return {
        "meta": meta,
        "realizado": round(caixa, 2),
        "pct": round(pct, 1),
        "falta": round(falta, 2),
        "dias_realizados": dias_realizados,
        "dias_total": dias_total,
        "dias_restantes": dias_restantes,
        "necessario_dia": round(necessario, 2),
        "ritmo_dia": round(ritmo, 2),
        "projecao": round(projecao, 2),
        "projecao_pct": round(projecao / max(meta, 1) * 100, 1),
    }


# ============================================================
# MAIN
# ============================================================

def hoje_ref() -> date:
    """Hoje como referência. Permite override via env DASHBOARD_HOJE=YYYY-MM-DD."""
    import os
    fx = os.environ.get("DASHBOARD_HOJE")
    if fx:
        return date.fromisoformat(fx)
    return date.today()


def semana_corrente(hoje: date) -> tuple[date, date]:
    """Segunda a domingo da semana em que HOJE cai."""
    seg = hoje - timedelta(days=hoje.weekday())
    dom = seg + timedelta(days=6)
    return seg, dom


def main():
    print("[build] Carregando dados...")
    h_ag, d_ag = carregar_agendamentos()
    h_tr, d_tr = carregar_transacoes()
    cad_map = carregar_clientes_cad()
    print(f"  {len(d_ag)} agendamentos · {len(d_tr)} transações · {len(cad_map)} clientes com data cad.")

    hoje = hoje_ref()
    ini_mes = date(hoje.year, hoje.month, 1)
    fim_mes = date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])
    seg, dom = semana_corrente(hoje)
    ini_ano = date(hoje.year, 1, 1)
    fim_ano = date(hoje.year, 12, 31)

    print(f"[build] Períodos: ano={ini_ano}..{fim_ano} · mês={ini_mes}..{fim_mes} · semana={seg}..{dom} · hoje={hoje}")

    # ------ ANUAL ------
    ag_anual = analisar_agendamentos(d_ag, h_ag, ini_ano, fim_ano)
    tr_anual = analisar_transacoes(d_tr, h_tr, ini_ano, fim_ano)
    ltv_ano = top_ltv(d_ag, h_ag, ini_ano, fim_ano, limite=15)

    # meses do ano para tabela mensal
    meses_ano = {}
    for m in range(1, 13):
        m_ini = date(hoje.year, m, 1)
        m_fim = date(hoje.year, m, monthrange(hoje.year, m)[1])
        m_ag = analisar_agendamentos(d_ag, h_ag, m_ini, m_fim)
        m_tr = analisar_transacoes(d_tr, h_tr, m_ini, m_fim)
        if m_ag["atend_fin"] > 0 or m_tr["n_trans"] > 0:
            meses_ano[f"{hoje.year}-{m:02d}"] = {
                "caixa": m_tr["caixa"],
                "receita_serv": m_ag["receita_serv"],
                "atend_fin": m_ag["atend_fin"],
                "clientes_unicos": m_ag["clientes_unicos"],
                "n_trans": m_tr["n_trans"],
                "ticket_trans": m_tr["ticket_trans"],
                "taxa_canc": m_ag["taxa_canc"],
                "categorias": m_tr["categorias"],
                "vendas_produtos": m_tr["categorias"]["produtos"]["v"],
                "dias_op": m_ag["dias_com_op"],
            }

    # ------ MENSAL ------
    ag_mes = analisar_agendamentos(d_ag, h_ag, ini_mes, fim_mes)
    tr_mes = analisar_transacoes(d_tr, h_tr, ini_mes, fim_mes)
    # novos vs recorr — usa só finalizados do mês
    i_st, i_dt = h_ag.index("status.nome"), h_ag.index("dataHoraInicio")
    fin_mes = [r for r in d_ag if r[i_st] == "Finalizado" and pd(r[i_dt]) and ini_mes <= pd(r[i_dt]).date() <= fim_mes]
    nvr = novos_vs_recorrentes(fin_mes, h_ag, cad_map, ini_mes)
    meta_mensal = calc_meta(tr_mes["caixa"], META_MENSAL, ag_mes["dias_com_op"], DIAS_OP_MES)

    # ------ SEMANAL ------
    ag_sem = analisar_agendamentos(d_ag, h_ag, seg, dom)
    tr_sem = analisar_transacoes(d_tr, h_tr, seg, dom)
    dias_op_sem = 6  # seg-sáb operacionais
    meta_sem_valor = META_MENSAL / DIAS_OP_MES * dias_op_sem
    meta_semanal = calc_meta(tr_sem["caixa"], round(meta_sem_valor, 2), ag_sem["dias_com_op"], dias_op_sem)

    # ------ DIÁRIO (hoje) ------
    ag_dia = analisar_agendamentos(d_ag, h_ag, hoje, hoje)
    tr_dia = analisar_transacoes(d_tr, h_tr, hoje, hoje)
    meta_dia_valor = META_MENSAL / DIAS_OP_MES
    meta_diaria = calc_meta(tr_dia["caixa"], round(meta_dia_valor, 2), 1, 1)

    # "em atendimento" hoje
    em_atend_hoje = sum(1 for r in d_ag
                        if r[h_ag.index("status.nome")] == "Em atendimento"
                        and pd(r[i_dt]) and pd(r[i_dt]).date() == hoje)

    # meta anual estimada (jul realizado + N meses × meta)
    dt_abertura = date(2026, 7, 23)
    meses_operacao = 5 if hoje.month <= 8 else max(1, 12 - hoje.month + 1)
    realizado_pre_ago = sum(m["caixa"] for k, m in meses_ano.items() if k < f"{hoje.year}-{max(hoje.month,1):02d}") if hoje.month > 1 else 0
    # simplificação: meta ano = realizado até ontem + META_MENSAL * meses restantes (inclui atual)
    meses_restantes = 12 - hoje.month + 1
    meta_ano = round(realizado_pre_ago + META_MENSAL * meses_restantes, 2)
    ano_realizado = tr_anual["caixa"]
    meta_anual_progresso = calc_meta(ano_realizado, meta_ano, hoje.month, 12) if meta_ano > 0 else {}

    # ------ HORA MEDIA por período ------
    def hora_media(hora_abs_list, dias_op):
        return [{"h": x["h"], "media": round(x["n"] / max(dias_op, 1), 2), "n_total": x["n"]}
                for x in hora_abs_list]

    payload = {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "hoje": hoje.isoformat(),
        "meta_mensal_valor": META_MENSAL,
        "dias_op_mes": DIAS_OP_MES,
        "cota_api_placeholder": {"total": 10000},  # atualizado por trinks_delta.py
        "abas": {
            "anual": {
                "kpis": {
                    "caixa": tr_anual["caixa"],
                    "receita_serv": ag_anual["receita_serv"],
                    "atend_fin": ag_anual["atend_fin"],
                    "n_trans": tr_anual["n_trans"],
                    "ticket_trans": tr_anual["ticket_trans"],
                    "taxa_canc": ag_anual["taxa_canc"],
                    "dias_op": ag_anual["dias_com_op"],
                    "clientes_unicos": ag_anual["clientes_unicos"],
                },
                "meta": meta_anual_progresso,
                "categorias": tr_anual["categorias"],
                "meses": meses_ano,
                "top_ltv": ltv_ano,
                "rentabilidade_hora": ag_anual["rentabilidade_hora"][:15],
                "hora_media": hora_media(tr_anual["hora_abs"], ag_anual["dias_com_op"]),
                "ranking_prof": ag_anual["ranking_prof"],
                "ranking_serv": ag_anual["ranking_serv"],
                "meios_pagamento": tr_anual["meios_pagamento"],
                "descontos": tr_anual["descontos"],
            },
            "mensal": {
                "kpis": {
                    "caixa": tr_mes["caixa"],
                    "receita_serv": ag_mes["receita_serv"],
                    "atend_fin": ag_mes["atend_fin"],
                    "n_trans": tr_mes["n_trans"],
                    "ticket_trans": tr_mes["ticket_trans"],
                    "taxa_canc": ag_mes["taxa_canc"],
                    "dias_op": ag_mes["dias_com_op"],
                    "clientes_unicos": ag_mes["clientes_unicos"],
                    "em_atendimento": em_atend_hoje,
                },
                "meta": meta_mensal,
                "categorias": tr_mes["categorias"],
                "novos_vs_recorr": nvr,
                "rentabilidade_hora": ag_mes["rentabilidade_hora"][:15],
                "por_dia_mes": ag_mes["por_dia_mes"],
                "hora_media": hora_media(tr_mes["hora_abs"], ag_mes["dias_com_op"]),
                "ranking_prof": ag_mes["ranking_prof"],
                "ranking_serv": ag_mes["ranking_serv"],
                "meios_pagamento": tr_mes["meios_pagamento"],
                "descontos": tr_mes["descontos"],
                "clientes_top": ag_mes["clientes_top"],
            },
            "semanal": {
                "kpis": {
                    "caixa": tr_sem["caixa"],
                    "atend_fin": ag_sem["atend_fin"],
                    "n_trans": tr_sem["n_trans"],
                    "ticket_trans": tr_sem["ticket_trans"],
                    "dias_op": ag_sem["dias_com_op"],
                    "clientes_unicos": ag_sem["clientes_unicos"],
                    "taxa_canc": ag_sem["taxa_canc"],
                    "periodo_ini": seg.isoformat(),
                    "periodo_fim": dom.isoformat(),
                },
                "meta": meta_semanal,
                "categorias": tr_sem["categorias"],
                "por_dow": ag_sem["por_dow"],
                "hora_media": hora_media(tr_sem["hora_abs"], ag_sem["dias_com_op"]),
                "ranking_prof": ag_sem["ranking_prof"],
                "ranking_serv": ag_sem["ranking_serv"],
                "meios_pagamento": tr_sem["meios_pagamento"],
            },
            "diario": {
                "kpis": {
                    "caixa": tr_dia["caixa"],
                    "receita_serv": ag_dia["receita_serv"],
                    "atend_fin": ag_dia["atend_fin"],
                    "n_trans": tr_dia["n_trans"],
                    "ticket_trans": tr_dia["ticket_trans"],
                    "clientes_unicos": ag_dia["clientes_unicos"],
                    "em_atendimento": em_atend_hoje,
                    "dia_semana": DOW_NOMES[hoje.weekday()],
                    "data": hoje.isoformat(),
                },
                "meta": meta_diaria,
                "categorias": tr_dia["categorias"],
                "hora_abs": tr_dia["hora_abs"],
                "ranking_prof": ag_dia["ranking_prof"],
                "ranking_serv": ag_dia["ranking_serv"],
                "meios_pagamento": tr_dia["meios_pagamento"],
            },
        },
    }

    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[build] OK · {OUT_JSON}")
    print(f"  Ano: caixa R$ {tr_anual['caixa']:.2f} · Mês: R$ {tr_mes['caixa']:.2f} · Semana: R$ {tr_sem['caixa']:.2f} · Hoje: R$ {tr_dia['caixa']:.2f}")


if __name__ == "__main__":
    main()
